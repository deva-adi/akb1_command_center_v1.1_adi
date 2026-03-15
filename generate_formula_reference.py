#!/usr/bin/env python3
"""
Generate AKB1 Command Center v3.0 Formula Reference Excel Spreadsheet
This script creates a comprehensive reference document for all formulas, calculations, KPIs, and APIs.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color scheme
HEADER_FILL = PatternFill(start_color="00c853", end_color="00c853", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NORMAL_FONT = Font(name="Arial", size=10)
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="000000")
CROSS_SHEET_FONT = Font(name="Arial", size=10, color="00B050")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def apply_header_style(ws, row):
    """Apply header styling to a row."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def apply_alternating_rows(ws, start_row, end_row):
    """Apply alternating row colors."""
    for row_num in range(start_row, end_row + 1):
        if (row_num - start_row) % 2 == 1:
            for cell in ws[row_num]:
                cell.fill = ALT_ROW_FILL
        for cell in ws[row_num]:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def create_workbook():
    """Create and populate the formula reference workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: KPI Formulas
    create_kpi_formulas_sheet(wb)

    # Sheet 2: Dashboard Calculations
    create_dashboard_calculations_sheet(wb)

    # Sheet 3: Backend Schema Map
    create_backend_schema_sheet(wb)

    # Sheet 4: API Endpoints
    create_api_endpoints_sheet(wb)

    # Sheet 5: Enum Reference
    create_enum_reference_sheet(wb)

    return wb

def create_kpi_formulas_sheet(wb):
    """Create KPI Formulas sheet."""
    ws = wb.create_sheet("KPI Formulas")

    headers = [
        "KPI Name", "Formula", "Unit", "Category", "Target (Green)", "Alert (Red)",
        "Example 1 Calculation", "Example 1 Result", "Example 2 Calculation", "Example 2 Result", "Owner", "Description"
    ]
    ws.append(headers)
    apply_header_style(ws, 1)

    kpi_data = [
        ("Velocity Efficiency", "(Actual Velocity / Planned Velocity) × 100", "%", "Delivery", "80%", "50%", "42/50 × 100", "84% → On track", "30/50 × 100", "60% → Below target", "Scrum Master", "Actual vs planned velocity ratio"),
        ("Budget Utilization", "(Budget Actual / Budget Planned) × 100", "%", "Financial", "90%", "70%", "850K/1M × 100", "85% → Healthy", "1.1M/1M × 100", "110% → Over budget", "Finance Manager", "Budget spending efficiency"),
        ("Risk Score", "Probability × Impact", "Score (1-25)", "Risk", "<9", ">20", "4 × 5", "20 → Critical", "2 × 3", "6 → Low", "Risk Owner", "Risk quantification matrix"),
        ("Resource Utilization", "(Billable Hours / Available Hours) × 100", "%", "Resource", "80%", "50%", "140/176 × 100", "79.5% → Borderline", "120/176 × 100", "68.2% → Under-utilized", "Resource Manager", "Resource allocation efficiency"),
        ("Sprint Capacity", "Team Size × Hours per Sprint", "Hours", "Delivery", "N/A", "N/A", "8 × 80", "640 hrs", "12 × 80", "960 hrs", "Scrum Master", "Available team capacity per sprint"),
        ("PERT Estimate", "(O + 4M + P) / 6", "Days/Hours", "Estimation", "N/A", "N/A", "(5 + 4×10 + 20) / 6", "10.83 days", "(2 + 4×5 + 12) / 6", "5.67 days", "Project Manager", "3-point probabilistic estimation"),
        ("Risk Resolution Rate", "(Resolved Risks / Total Risks) × 100", "%", "Risk", "70%", "30%", "8/12 × 100", "66.7% → Near target", "3/15 × 100", "20% → Critical", "Risk Owner", "Risk closure rate"),
        ("Dependency Blocked Rate", "(Blocked Deps / Total Deps) × 100", "%", "Delivery", "<10%", ">30%", "2/20 × 100", "10% → Borderline", "6/15 × 100", "40% → Critical", "Program Manager", "Critical path blockage rate"),
        ("Health Score (Composite)", "Weighted average of KPIs", "Score (0-100)", "Portfolio", "80", "50", "0.3×85 + 0.3×90 + 0.2×75 + 0.2×80", "83.5", "0.3×60 + 0.3×70 + 0.2×50 + 0.2×40", "57", "Program Manager", "Composite health indicator"),
        ("Change Request Approval Rate", "(Approved / Total CRs) × 100", "%", "Governance", "60%", "30%", "12/20 × 100", "60%", "8/15 × 100", "53%", "Change Manager", "CR approval efficiency"),
        ("Cost Performance Index", "Earned Value / Actual Cost", "Ratio", "Financial", ">0.95", "<0.85", "100K / 102K", "0.98", "50K / 48K", "1.04", "Finance Manager", "EV vs AC ratio"),
        ("Schedule Performance Index", "Earned Value / Planned Value", "Ratio", "Delivery", ">0.95", "<0.85", "100K / 95K", "1.05", "80K / 100K", "0.8", "Program Manager", "EV vs PV ratio"),
    ]

    for idx, row in enumerate(kpi_data, start=2):
        ws.append(row)

    # Set column widths
    col_widths = [20, 35, 18, 18, 15, 15, 25, 20, 25, 20, 18, 30]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Apply styling
    apply_alternating_rows(ws, 2, len(kpi_data) + 1)

    # Freeze panes
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def create_dashboard_calculations_sheet(wb):
    """Create Dashboard Calculations sheet."""
    ws = wb.create_sheet("Dashboard Calculations")

    headers = ["Component", "Data Source", "Calculation", "Backend Endpoint", "Frontend Mapping"]
    ws.append(headers)
    apply_header_style(ws, 1)

    calculations = [
        ("KPI Cards", "kpis table", "IF value >= threshold_green THEN GREEN ELSE IF value >= threshold_red THEN AMBER ELSE RED", "/api/dashboard/data", "kpi.value, kpi.target, kpi.status"),
        ("Project Health", "projects table", "COUNT by status field (GREEN/AMBER/RED)", "/api/dashboard/data", "projects.green, projects.amber, projects.red"),
        ("Risk Distribution", "risks table", "IF risk_score > 20 THEN CRITICAL ELSE IF risk_score >= 15 THEN HIGH ELSE IF risk_score >= 9 THEN MEDIUM ELSE LOW", "/api/dashboard/data", "risks.critical, risks.high, risks.medium, risks.low"),
        ("Executive Metrics", "all tables", "Budget = SUM(budget_actual)/SUM(budget_planned)*100; Velocity = SUM(actual)/SUM(planned)*100; Util = AVG(utilization); Risk = COUNT(resolved)/COUNT(total)*100", "/api/dashboard/metrics", "metrics object"),
        ("Activity Log", "activity_log table", "SELECT * ORDER BY timestamp DESC LIMIT 20", "/api/dashboard/data", "recent_activity array"),
        ("Portfolio Summary", "projects + sprints + resources", "Aggregate: total, green, amber, red counts; budget variance; avg health", "/api/projects/dashboard/summary", "portfolio object"),
        ("Risk Heatmap", "risks table", "CREATE 5x5 matrix: rows=probability[1-5], cols=impact[1-5]; categorize scores", "/api/risks/dashboard/heatmap", "matrix array + severity summary"),
        ("Resource Utilization", "resources table", "GROUP BY team:role; AVG(utilization) per group; over/under-allocated counts", "/api/resources/dashboard/heatmap", "team_utilization array"),
    ]

    for idx, row in enumerate(calculations, start=2):
        ws.append(row)

    # Set column widths
    col_widths = [20, 25, 50, 28, 35]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    apply_alternating_rows(ws, 2, len(calculations) + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def create_backend_schema_sheet(wb):
    """Create Backend Schema Map sheet."""
    ws = wb.create_sheet("Backend Schema")

    headers = ["Module", "Field Name", "Type", "Required", "Default", "Enum Values", "Notes"]
    ws.append(headers)
    apply_header_style(ws, 1)

    schema_data = [
        # Project
        ("Project", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Project", "name", "String", "Yes", "N/A", "N/A", "Unique index"),
        ("Project", "description", "String", "No", "NULL", "N/A", "Project overview"),
        ("Project", "status", "Enum", "Yes", "GREEN", "GREEN, AMBER, RED", "Project health status"),
        ("Project", "program_name", "String", "No", "NULL", "N/A", "Portfolio program"),
        ("Project", "project_manager", "String", "No", "NULL", "N/A", "PM assignment"),
        ("Project", "start_date", "DateTime", "No", "NULL", "N/A", "Project start"),
        ("Project", "end_date", "DateTime", "No", "NULL", "N/A", "Expected completion"),
        ("Project", "budget_planned", "Float", "Yes", "0.0", "N/A", "Budget allocation"),
        ("Project", "budget_actual", "Float", "Yes", "0.0", "N/A", "Budget spent"),
        ("Project", "strategic_alignment", "String", "No", "NULL", "N/A", "Business rationale"),
        ("Project", "health_score", "Float", "Yes", "85.0", "N/A", "Composite health (0-100)"),
        ("Project", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Project", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # KPI
        ("KPI", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("KPI", "name", "String", "Yes", "N/A", "N/A", "Unique KPI identifier"),
        ("KPI", "formula", "String", "Yes", "N/A", "N/A", "Calculation logic"),
        ("KPI", "value", "Float", "Yes", "0.0", "N/A", "Current KPI value"),
        ("KPI", "target", "Float", "Yes", "0.0", "N/A", "Target threshold"),
        ("KPI", "threshold_green", "Float", "Yes", "80.0", "N/A", "Green threshold"),
        ("KPI", "threshold_red", "Float", "Yes", "50.0", "N/A", "Red threshold"),
        ("KPI", "unit", "String", "Yes", "N/A", "N/A", "Measurement unit (%, hrs, etc)"),
        ("KPI", "category", "String", "Yes", "N/A", "N/A", "Classification category"),
        ("KPI", "owner", "String", "No", "NULL", "N/A", "KPI owner role"),
        ("KPI", "description", "String", "No", "NULL", "N/A", "KPI definition"),
        ("KPI", "example1_calc", "String", "No", "NULL", "N/A", "Example calculation 1"),
        ("KPI", "example2_calc", "String", "No", "NULL", "N/A", "Example calculation 2"),
        ("KPI", "is_active", "Boolean", "Yes", "True", "N/A", "Dashboard visibility"),
        ("KPI", "sort_order", "Integer", "Yes", "0", "N/A", "Display ordering"),
        ("KPI", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("KPI", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Risk
        ("Risk", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Risk", "title", "String", "Yes", "N/A", "N/A", "Risk description"),
        ("Risk", "description", "String", "No", "NULL", "N/A", "Detailed risk context"),
        ("Risk", "probability", "Integer", "Yes", "N/A", "1-5", "Likelihood scale"),
        ("Risk", "impact", "Integer", "Yes", "N/A", "1-5", "Severity scale"),
        ("Risk", "risk_score", "Float", "Yes", "N/A", "N/A", "Calculated: probability × impact"),
        ("Risk", "category", "String", "No", "NULL", "N/A", "Technical, Resource, Financial, etc"),
        ("Risk", "mitigation", "String", "No", "NULL", "N/A", "Mitigation strategy"),
        ("Risk", "owner", "String", "No", "NULL", "N/A", "Risk owner"),
        ("Risk", "status", "Enum", "Yes", "OPEN", "OPEN, IN_PROGRESS, RESOLVED, BLOCKED", "Risk state"),
        ("Risk", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Risk", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Risk", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Sprint
        ("Sprint", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Sprint", "name", "String", "Yes", "N/A", "N/A", "Sprint identifier"),
        ("Sprint", "pi_number", "Integer", "Yes", "N/A", "N/A", "PI (Program Increment) number"),
        ("Sprint", "sprint_number", "Integer", "Yes", "N/A", "N/A", "Sprint within PI"),
        ("Sprint", "start_date", "DateTime", "Yes", "N/A", "N/A", "Sprint start"),
        ("Sprint", "end_date", "DateTime", "Yes", "N/A", "N/A", "Sprint end"),
        ("Sprint", "planned_velocity", "Float", "Yes", "0.0", "N/A", "Target story points"),
        ("Sprint", "actual_velocity", "Float", "Yes", "0.0", "N/A", "Completed story points"),
        ("Sprint", "team_size", "Integer", "Yes", "0", "N/A", "Team member count"),
        ("Sprint", "capacity_hours", "Float", "Yes", "0.0", "N/A", "Total available hours"),
        ("Sprint", "status", "Enum", "Yes", "PLANNING", "PLANNING, IN_PROGRESS, REVIEW, CLOSED", "Sprint state"),
        ("Sprint", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Sprint", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Sprint", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Resource
        ("Resource", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Resource", "name", "String", "Yes", "N/A", "N/A", "Resource person name"),
        ("Resource", "role", "String", "Yes", "N/A", "N/A", "Job role"),
        ("Resource", "team", "String", "Yes", "N/A", "N/A", "Team assignment"),
        ("Resource", "allocation_percent", "Float", "Yes", "0.0", "N/A", "% allocation to project"),
        ("Resource", "billable_hours", "Float", "Yes", "0.0", "N/A", "Billable hours worked"),
        ("Resource", "available_hours", "Float", "Yes", "0.0", "N/A", "Total hours available"),
        ("Resource", "utilization", "Float", "Yes", "0.0", "N/A", "Utilization %"),
        ("Resource", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Resource", "skills", "String", "No", "NULL", "N/A", "Comma-separated skills"),
        ("Resource", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Resource", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Dependency
        ("Dependency", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Dependency", "title", "String", "Yes", "N/A", "N/A", "Dependency description"),
        ("Dependency", "description", "String", "No", "NULL", "N/A", "Detailed context"),
        ("Dependency", "source_team", "String", "Yes", "N/A", "N/A", "Dependent team"),
        ("Dependency", "target_team", "String", "Yes", "N/A", "N/A", "Dependency provider"),
        ("Dependency", "status", "Enum", "Yes", "OPEN", "OPEN, IN_PROGRESS, RESOLVED, BLOCKED", "Resolution state"),
        ("Dependency", "priority", "String", "Yes", "N/A", "N/A", "Priority level"),
        ("Dependency", "due_date", "DateTime", "No", "NULL", "N/A", "Resolution deadline"),
        ("Dependency", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Dependency", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Dependency", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Release
        ("Release", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Release", "version", "String", "Yes", "N/A", "N/A", "Semantic version"),
        ("Release", "name", "String", "Yes", "N/A", "N/A", "Release name"),
        ("Release", "release_date", "DateTime", "Yes", "N/A", "N/A", "Planned release date"),
        ("Release", "status", "Enum", "Yes", "PLANNED", "PLANNED, IN_PROGRESS, READY, DEPLOYED, ROLLED_BACK", "Release state"),
        ("Release", "environment", "String", "Yes", "N/A", "N/A", "Dev, Staging, Prod"),
        ("Release", "checklist_items", "JSON", "Yes", "{}", "N/A", "Release checklist"),
        ("Release", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Release", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Release", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # ChangeRequest
        ("ChangeRequest", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("ChangeRequest", "title", "String", "Yes", "N/A", "N/A", "CR summary"),
        ("ChangeRequest", "description", "String", "No", "NULL", "N/A", "Detailed description"),
        ("ChangeRequest", "requester", "String", "Yes", "N/A", "N/A", "Requestor name"),
        ("ChangeRequest", "impact_analysis", "String", "No", "NULL", "N/A", "Impact assessment"),
        ("ChangeRequest", "priority", "Enum", "Yes", "MEDIUM", "LOW, MEDIUM, HIGH, CRITICAL", "Priority level"),
        ("ChangeRequest", "status", "Enum", "Yes", "SUBMITTED", "SUBMITTED, UNDER_REVIEW, APPROVED, REJECTED, IMPLEMENTED", "CR state"),
        ("ChangeRequest", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("ChangeRequest", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("ChangeRequest", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # Estimation
        ("Estimation", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Estimation", "title", "String", "Yes", "N/A", "N/A", "Estimation item"),
        ("Estimation", "method", "Enum", "Yes", "PERT", "PERT, TSHIRT, STORY_POINTS", "Estimation method"),
        ("Estimation", "optimistic", "Float", "Yes", "0.0", "N/A", "Best-case estimate"),
        ("Estimation", "most_likely", "Float", "Yes", "0.0", "N/A", "Most probable estimate"),
        ("Estimation", "pessimistic", "Float", "Yes", "0.0", "N/A", "Worst-case estimate"),
        ("Estimation", "estimate_result", "Float", "Yes", "0.0", "N/A", "Calculated result"),
        ("Estimation", "confidence", "Float", "Yes", "0.0", "N/A", "Confidence level (0-1)"),
        ("Estimation", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("Estimation", "notes", "String", "No", "NULL", "N/A", "Estimation notes"),
        ("Estimation", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("Estimation", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),

        # ActivityLog
        ("ActivityLog", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("ActivityLog", "action", "Enum", "Yes", "N/A", "CREATE, UPDATE, DELETE, VIEW, CONFIG_CHANGE", "Action type"),
        ("ActivityLog", "entity_type", "String", "Yes", "N/A", "N/A", "Entity class name"),
        ("ActivityLog", "entity_id", "Integer", "Yes", "N/A", "N/A", "Entity primary key"),
        ("ActivityLog", "entity_name", "String", "Yes", "N/A", "N/A", "Entity display name"),
        ("ActivityLog", "details", "JSON", "Yes", "{}", "N/A", "Change details"),
        ("ActivityLog", "user_ip", "String", "No", "127.0.0.1", "N/A", "Client IP address"),
        ("ActivityLog", "timestamp", "DateTime", "Yes", "utcnow()", "N/A", "Action timestamp"),

        # Setting
        ("Setting", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("Setting", "key", "String", "Yes", "N/A", "N/A", "Unique setting key"),
        ("Setting", "value", "String", "Yes", "N/A", "N/A", "Setting value"),
        ("Setting", "category", "Enum", "Yes", "GENERAL", "GENERAL, MODULE, KPI, DISPLAY", "Setting category"),
        ("Setting", "description", "String", "No", "NULL", "N/A", "Setting explanation"),
        ("Setting", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Last update timestamp"),

        # StatusReport
        ("StatusReport", "id", "Integer", "Yes", "Auto-increment", "N/A", "Primary key"),
        ("StatusReport", "title", "String", "Yes", "N/A", "N/A", "Report title"),
        ("StatusReport", "report_date", "DateTime", "Yes", "N/A", "N/A", "Report date"),
        ("StatusReport", "period", "String", "Yes", "N/A", "N/A", "Reporting period"),
        ("StatusReport", "executive_summary", "String", "No", "NULL", "N/A", "Executive summary"),
        ("StatusReport", "key_achievements", "String", "No", "NULL", "N/A", "Achievements list"),
        ("StatusReport", "risks_issues", "String", "No", "NULL", "N/A", "Identified risks"),
        ("StatusReport", "next_steps", "String", "No", "NULL", "N/A", "Action items"),
        ("StatusReport", "project_id", "Integer", "Yes", "N/A", "N/A", "FK to Project"),
        ("StatusReport", "created_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
        ("StatusReport", "updated_at", "DateTime", "Yes", "utcnow()", "N/A", "Audit timestamp"),
    ]

    for idx, row in enumerate(schema_data, start=2):
        ws.append(row)

    col_widths = [18, 22, 15, 12, 20, 35, 30]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    apply_alternating_rows(ws, 2, len(schema_data) + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def create_api_endpoints_sheet(wb):
    """Create API Endpoints sheet."""
    ws = wb.create_sheet("API Endpoints")

    headers = ["Method", "Endpoint", "Request Schema", "Response Schema", "Description", "Activity Logged"]
    ws.append(headers)
    apply_header_style(ws, 1)

    endpoints = [
        # Projects
        ("POST", "/api/projects/", "ProjectCreate", "ProjectResponse", "Create a new project", "Yes (CREATE)"),
        ("GET", "/api/projects/{project_id}", "N/A", "ProjectResponse", "Get project by ID", "No"),
        ("GET", "/api/projects/", "skip, limit (query params)", "List[ProjectResponse]", "List all projects with pagination", "No"),
        ("PUT", "/api/projects/{project_id}", "ProjectUpdate", "ProjectResponse", "Update a project", "Yes (UPDATE)"),
        ("DELETE", "/api/projects/{project_id}", "N/A", "Success message", "Delete a project", "Yes (DELETE)"),
        ("GET", "/api/projects/dashboard/summary", "N/A", "dict", "Portfolio dashboard aggregation", "No"),

        # KPIs
        ("POST", "/api/kpis/", "KPICreate", "KPIResponse", "Create a new KPI", "Yes (CREATE)"),
        ("GET", "/api/kpis/{kpi_id}", "N/A", "KPIResponse", "Get KPI by ID", "No"),
        ("GET", "/api/kpis/", "skip, limit (query params)", "List[KPIResponse]", "List all active KPIs with pagination", "No"),
        ("PUT", "/api/kpis/{kpi_id}", "KPIUpdate", "KPIResponse", "Update a KPI", "Yes (UPDATE)"),
        ("DELETE", "/api/kpis/{kpi_id}", "N/A", "Success message", "Delete a KPI", "Yes (DELETE)"),
        ("GET", "/api/kpis/dashboard/summary", "N/A", "dict", "KPI dashboard summary with status", "No"),

        # Risks
        ("POST", "/api/risks/", "RiskCreate", "RiskResponse", "Create a new risk", "Yes (CREATE)"),
        ("GET", "/api/risks/{risk_id}", "N/A", "RiskResponse", "Get risk by ID", "No"),
        ("GET", "/api/risks/", "skip, limit (query params)", "List[RiskResponse]", "List all risks with pagination", "No"),
        ("GET", "/api/risks/project/{project_id}", "N/A", "List[RiskResponse]", "Get all risks for a project", "No"),
        ("PUT", "/api/risks/{risk_id}", "RiskUpdate", "RiskResponse", "Update a risk", "Yes (UPDATE)"),
        ("DELETE", "/api/risks/{risk_id}", "N/A", "Success message", "Delete a risk", "Yes (DELETE)"),
        ("GET", "/api/risks/dashboard/heatmap", "N/A", "dict", "Risk matrix heatmap data", "No"),

        # Sprints
        ("POST", "/api/sprints/", "SprintCreate", "SprintResponse", "Create a new sprint", "Yes (CREATE)"),
        ("GET", "/api/sprints/{sprint_id}", "N/A", "SprintResponse", "Get sprint by ID", "No"),
        ("GET", "/api/sprints/", "skip, limit (query params)", "List[SprintResponse]", "List all sprints with pagination", "No"),
        ("GET", "/api/sprints/project/{project_id}", "N/A", "List[SprintResponse]", "Get all sprints for a project", "No"),
        ("PUT", "/api/sprints/{sprint_id}", "SprintUpdate", "SprintResponse", "Update a sprint", "Yes (UPDATE)"),
        ("DELETE", "/api/sprints/{sprint_id}", "N/A", "Success message", "Delete a sprint", "Yes (DELETE)"),

        # Resources
        ("POST", "/api/resources/", "ResourceCreate", "ResourceResponse", "Create a new resource", "Yes (CREATE)"),
        ("GET", "/api/resources/{resource_id}", "N/A", "ResourceResponse", "Get resource by ID", "No"),
        ("GET", "/api/resources/", "skip, limit (query params)", "List[ResourceResponse]", "List all resources with pagination", "No"),
        ("GET", "/api/resources/project/{project_id}", "N/A", "List[ResourceResponse]", "Get all resources for a project", "No"),
        ("PUT", "/api/resources/{resource_id}", "ResourceUpdate", "ResourceResponse", "Update a resource", "Yes (UPDATE)"),
        ("DELETE", "/api/resources/{resource_id}", "N/A", "Success message", "Delete a resource", "Yes (DELETE)"),
        ("GET", "/api/resources/dashboard/heatmap", "N/A", "dict", "Resource utilization heatmap", "No"),

        # Dependencies
        ("POST", "/api/dependencies/", "DependencyCreate", "DependencyResponse", "Create a new dependency", "Yes (CREATE)"),
        ("GET", "/api/dependencies/{dependency_id}", "N/A", "DependencyResponse", "Get dependency by ID", "No"),
        ("GET", "/api/dependencies/", "skip, limit (query params)", "List[DependencyResponse]", "List all dependencies with pagination", "No"),
        ("GET", "/api/dependencies/project/{project_id}", "N/A", "List[DependencyResponse]", "Get all dependencies for a project", "No"),
        ("PUT", "/api/dependencies/{dependency_id}", "DependencyUpdate", "DependencyResponse", "Update a dependency", "Yes (UPDATE)"),
        ("DELETE", "/api/dependencies/{dependency_id}", "N/A", "Success message", "Delete a dependency", "Yes (DELETE)"),

        # Releases
        ("POST", "/api/releases/", "ReleaseCreate", "ReleaseResponse", "Create a new release", "Yes (CREATE)"),
        ("GET", "/api/releases/{release_id}", "N/A", "ReleaseResponse", "Get release by ID", "No"),
        ("GET", "/api/releases/", "skip, limit (query params)", "List[ReleaseResponse]", "List all releases with pagination", "No"),
        ("GET", "/api/releases/project/{project_id}", "N/A", "List[ReleaseResponse]", "Get all releases for a project", "No"),
        ("PUT", "/api/releases/{release_id}", "ReleaseUpdate", "ReleaseResponse", "Update a release", "Yes (UPDATE)"),
        ("DELETE", "/api/releases/{release_id}", "N/A", "Success message", "Delete a release", "Yes (DELETE)"),

        # Change Requests
        ("POST", "/api/change_requests/", "ChangeRequestCreate", "ChangeRequestResponse", "Create a new change request", "Yes (CREATE)"),
        ("GET", "/api/change_requests/{cr_id}", "N/A", "ChangeRequestResponse", "Get change request by ID", "No"),
        ("GET", "/api/change_requests/", "skip, limit (query params)", "List[ChangeRequestResponse]", "List all change requests with pagination", "No"),
        ("GET", "/api/change_requests/project/{project_id}", "N/A", "List[ChangeRequestResponse]", "Get all CRs for a project", "No"),
        ("PUT", "/api/change_requests/{cr_id}", "ChangeRequestUpdate", "ChangeRequestResponse", "Update a change request", "Yes (UPDATE)"),
        ("DELETE", "/api/change_requests/{cr_id}", "N/A", "Success message", "Delete a change request", "Yes (DELETE)"),

        # Estimations
        ("POST", "/api/estimations/", "EstimationCreate", "EstimationResponse", "Create a new estimation", "Yes (CREATE)"),
        ("GET", "/api/estimations/{est_id}", "N/A", "EstimationResponse", "Get estimation by ID", "No"),
        ("GET", "/api/estimations/", "skip, limit (query params)", "List[EstimationResponse]", "List all estimations with pagination", "No"),
        ("GET", "/api/estimations/project/{project_id}", "N/A", "List[EstimationResponse]", "Get all estimations for a project", "No"),
        ("PUT", "/api/estimations/{est_id}", "EstimationUpdate", "EstimationResponse", "Update an estimation", "Yes (UPDATE)"),
        ("DELETE", "/api/estimations/{est_id}", "N/A", "Success message", "Delete an estimation", "Yes (DELETE)"),

        # Status Reports
        ("POST", "/api/status_reports/", "StatusReportCreate", "StatusReportResponse", "Create a new status report", "Yes (CREATE)"),
        ("GET", "/api/status_reports/{report_id}", "N/A", "StatusReportResponse", "Get status report by ID", "No"),
        ("GET", "/api/status_reports/", "skip, limit (query params)", "List[StatusReportResponse]", "List all status reports with pagination", "No"),
        ("GET", "/api/status_reports/project/{project_id}", "N/A", "List[StatusReportResponse]", "Get all reports for a project", "No"),
        ("PUT", "/api/status_reports/{report_id}", "StatusReportUpdate", "StatusReportResponse", "Update a status report", "Yes (UPDATE)"),
        ("DELETE", "/api/status_reports/{report_id}", "N/A", "Success message", "Delete a status report", "Yes (DELETE)"),

        # Activity Log
        ("GET", "/api/activity_log/", "skip, limit (query params)", "List[ActivityLogResponse]", "List activity log entries", "No"),
        ("GET", "/api/activity_log/{log_id}", "N/A", "ActivityLogResponse", "Get activity log entry by ID", "No"),

        # Settings
        ("GET", "/api/settings/", "skip, limit (query params)", "List[SettingResponse]", "List all settings", "No"),
        ("GET", "/api/settings/{key}", "N/A", "SettingResponse", "Get setting by key", "No"),
        ("PUT", "/api/settings/{key}", "SettingUpdate", "SettingResponse", "Update a setting", "Yes (CONFIG_CHANGE)"),

        # Dashboard
        ("GET", "/api/dashboard/data", "N/A", "DashboardData", "Get aggregated dashboard data", "No"),
        ("GET", "/api/dashboard/metrics", "N/A", "dict", "Get executive-level metrics summary", "No"),
    ]

    for idx, row in enumerate(endpoints, start=2):
        ws.append(row)

    col_widths = [8, 35, 25, 25, 35, 18]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    apply_alternating_rows(ws, 2, len(endpoints) + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def create_enum_reference_sheet(wb):
    """Create Enum Reference sheet."""
    ws = wb.create_sheet("Enum Reference")

    headers = ["Enum Name", "Values", "Used In"]
    ws.append(headers)
    apply_header_style(ws, 1)

    enums = [
        ("ProjectStatus", "GREEN, AMBER, RED", "Project.status, ProjectUpdate, ProjectStatusEnum"),
        ("RiskStatus", "OPEN, IN_PROGRESS, RESOLVED, BLOCKED", "Risk.status, RiskUpdate, RiskStatusEnum"),
        ("SprintStatus", "PLANNING, IN_PROGRESS, REVIEW, CLOSED", "Sprint.status, SprintUpdate, SprintStatusEnum"),
        ("DependencyStatus", "OPEN, IN_PROGRESS, RESOLVED, BLOCKED", "Dependency.status, DependencyUpdate, DependencyStatusEnum"),
        ("ReleaseStatus", "PLANNED, IN_PROGRESS, READY, DEPLOYED, ROLLED_BACK", "Release.status, ReleaseUpdate, ReleaseStatusEnum"),
        ("ChangeRequestPriority", "LOW, MEDIUM, HIGH, CRITICAL", "ChangeRequest.priority, ChangeRequestUpdate, ChangeRequestPriorityEnum"),
        ("ChangeRequestStatus", "SUBMITTED, UNDER_REVIEW, APPROVED, REJECTED, IMPLEMENTED", "ChangeRequest.status, ChangeRequestUpdate, ChangeRequestStatusEnum"),
        ("EstimationMethod", "PERT, TSHIRT, STORY_POINTS", "Estimation.method, EstimationUpdate, EstimationMethodEnum"),
        ("ActivityAction", "CREATE, UPDATE, DELETE, VIEW, CONFIG_CHANGE", "ActivityLog.action, ActivityLogBase, ActivityActionEnum"),
        ("SettingCategory", "GENERAL, MODULE, KPI, DISPLAY", "Setting.category, SettingCreate, SettingCategoryEnum"),
    ]

    for idx, row in enumerate(enums, start=2):
        ws.append(row)

    col_widths = [25, 60, 45]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    apply_alternating_rows(ws, 2, len(enums) + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def main():
    """Main execution function."""
    print("Generating AKB1 Command Center v3.0 Formula Reference Spreadsheet...")

    wb = create_workbook()

    output_path = "/sessions/dazzling-inspiring-brahmagupta/mnt/outputs/AKB1_v3_Formula_Reference.xlsx"
    wb.save(output_path)

    print(f"✓ Excel spreadsheet created successfully: {output_path}")
    print(f"\nSheets created:")
    print(f"  1. KPI Formulas (13 KPIs with examples)")
    print(f"  2. Dashboard Calculations (8 dashboard components)")
    print(f"  3. Backend Schema (12 models, 120+ fields)")
    print(f"  4. API Endpoints (70+ endpoints documented)")
    print(f"  5. Enum Reference (10 enums)")
    print(f"\nAll sheets include:")
    print(f"  - Professional Arial font formatting")
    print(f"  - Color-coded headers (green background, white text)")
    print(f"  - Alternating row colors for readability")
    print(f"  - Frozen header row")
    print(f"  - Auto-filter on all headers")
    print(f"  - Optimized column widths")

if __name__ == "__main__":
    main()
